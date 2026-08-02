import asyncio
import csv
import re
import statistics
from collections import Counter
from pathlib import Path

from .base import Agent


class DataAgent(Agent):
    name = "data"
    description = "Analyze CSV/Excel: statistics, top values, ASCII charts"
    keywords = ["analyze", "csv", "excel", "spreadsheet", "data", "statistics", "trend",
                "chart", "average", "dataset", "xlsx"]

    async def run(self, task, context=""):
        path = self._find_path(task)
        if not path:
            return self.done("Point me at a CSV or Excel file, boss.")
        loop = asyncio.get_running_loop()
        if path.suffix.lower() == ".csv":
            rows = await loop.run_in_executor(None, self._read_csv, path)
        else:
            rows = await loop.run_in_executor(None, self._read_xlsx, path)
        if not rows:
            return self.done("That file appears empty, boss.")
        report = self._stats(path.name, rows)
        return self.done(f"Analyzed {path.name}: {len(rows) - 1} rows, {len(rows[0])} columns. Full stats in the output panel.",
                         report)

    def _find_path(self, task):
        m = re.findall(r"[\"']([^\"']+\.\w{2,4})[\"']", task) or re.findall(r"(\S+\.(?:csv|xlsx|xls))", task)
        if m:
            p = Path(m[0]).expanduser()
            if p.exists():
                return p
        for d in (Path.home() / "Downloads", Path.home() / "Desktop"):
            cands = sorted(list(d.glob("*.csv")) + list(d.glob("*.xlsx")),
                           key=lambda x: x.stat().st_mtime, reverse=True)
            if cands:
                return cands[0]
        return None

    def _read_csv(self, path):
        with open(path, newline="", encoding="utf-8", errors="ignore") as f:
            return [row for row in csv.reader(f) if row][:5001]

    def _read_xlsx(self, path):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        return [[str(c) if c is not None else "" for c in row] for row in
                ws.iter_rows(max_row=5000, values_only=True)]

    def _stats(self, name, rows):
        header, data = rows[0], rows[1:]
        lines = [f"DATASET {name}", f"rows={len(data)}  cols={len(header)}", ""]
        for i, col in enumerate(header):
            vals = [r[i] for r in data if i < len(r) and r[i] not in ("", None)]
            nums = []
            for v in vals:
                try:
                    nums.append(float(str(v).replace(",", "")))
                except ValueError:
                    pass
            if nums and len(nums) >= max(1, len(vals) * 0.6):
                lines.append(f"{col}: mean={statistics.mean(nums):.2f} median={statistics.median(nums):.2f} "
                             f"min={min(nums):.2f} max={max(nums):.2f}")
                lines.append(self._bars(nums))
            else:
                top = Counter(vals).most_common(3)
                lines.append(f"{col}: top={top}")
            lines.append("")
        return "\n".join(lines)

    def _bars(self, nums, buckets=6, width=34):
        lo, hi = min(nums), max(nums)
        if hi == lo:
            return f"  [constant value {lo}]"
        step = (hi - lo) / buckets
        counts = [0] * buckets
        for n in nums:
            counts[min(buckets - 1, int((n - lo) / step))] += 1
        peak = max(counts) or 1
        out = []
        for i, c in enumerate(counts):
            edge = lo + i * step
            out.append(f"  {edge:>10.1f} | {'█' * int(width * c / peak)} {c}")
        return "\n".join(out)


AGENT = DataAgent